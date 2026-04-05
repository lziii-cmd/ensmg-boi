import uuid
import openpyxl
from django.utils import timezone
from rest_framework import status, generics
from rest_framework.decorators import api_view, permission_classes
from rest_framework.permissions import AllowAny, IsAuthenticated
from rest_framework.response import Response
from rest_framework.views import APIView
from rest_framework_simplejwt.tokens import RefreshToken

from .models import User, MemberImport, AuditLog
from .permissions import IsAdmin, IsResponsableOrAdmin, IsAuditor
from .serializers import (
    UserSerializer, LoginSerializer, SetPasswordSerializer,
    ChangePasswordSerializer, MemberImportSerializer, AuditLogSerializer,
)
from .tasks import send_invitation_email, send_password_reset_email


def get_client_ip(request):
    x_forwarded = request.META.get("HTTP_X_FORWARDED_FOR")
    if x_forwarded:
        return x_forwarded.split(",")[0].strip()
    return request.META.get("REMOTE_ADDR", "")


def log_action(user, action, target_type="", target_id="", target_repr="", details=None, ip=""):
    """Helper pour créer un log d'audit."""
    AuditLog.objects.create(
        user=user,
        action=action,
        target_type=target_type,
        target_id=str(target_id) if target_id else "",
        target_repr=target_repr,
        details=details or {},
        ip_address=ip,
    )


class LoginView(APIView):
    permission_classes = [AllowAny]

    def post(self, request):
        serializer = LoginSerializer(data=request.data)
        if not serializer.is_valid():
            error = serializer.errors.get("non_field_errors", ["Identifiants invalides."])
            return Response(
                {"detail": error[0] if isinstance(error, list) else str(error)},
                status=status.HTTP_400_BAD_REQUEST,
            )

        user = serializer.validated_data["user"]
        refresh = RefreshToken.for_user(user)

        log_action(
            user=user,
            action=AuditLog.LOGIN,
            target_repr=user.email,
            ip=get_client_ip(request),
        )

        return Response({
            "access": str(refresh.access_token),
            "refresh": str(refresh),
            "user": UserSerializer(user).data,
        })


class LogoutView(APIView):
    def post(self, request):
        log_action(
            user=request.user,
            action=AuditLog.LOGOUT,
            target_repr=request.user.email,
            ip=get_client_ip(request),
        )
        try:
            refresh_token = request.data.get("refresh")
            token = RefreshToken(refresh_token)
            token.blacklist()
        except Exception:
            pass
        return Response(status=status.HTTP_204_NO_CONTENT)


class SetPasswordView(APIView):
    permission_classes = [AllowAny]

    def post(self, request):
        serializer = SetPasswordSerializer(data=request.data)
        if not serializer.is_valid():
            return Response(serializer.errors, status=status.HTTP_400_BAD_REQUEST)

        user = serializer.validated_data["user"]
        user.set_password(serializer.validated_data["password"])
        user.password_set = True
        user.invitation_token = uuid.uuid4()
        user.save()
        return Response({"detail": "Mot de passe défini avec succès."})


class PasswordResetRequestView(APIView):
    permission_classes = [AllowAny]

    def post(self, request):
        email = request.data.get("email", "").lower().strip()
        try:
            user = User.objects.get(email=email, is_active=True)
            user.invitation_token = uuid.uuid4()
            user.save(update_fields=["invitation_token"])
            send_password_reset_email.delay(str(user.id))
        except User.DoesNotExist:
            pass
        return Response({"detail": "Si cet email existe, un lien de réinitialisation a été envoyé."})


class MeView(generics.RetrieveUpdateAPIView):
    serializer_class = UserSerializer

    def get_object(self):
        return self.request.user

    def update(self, request, *args, **kwargs):
        kwargs["partial"] = True
        return super().update(request, *args, **kwargs)


class ChangePasswordView(APIView):
    def post(self, request):
        serializer = ChangePasswordSerializer(data=request.data)
        if not serializer.is_valid():
            return Response(serializer.errors, status=status.HTTP_400_BAD_REQUEST)

        user = request.user
        if not user.check_password(serializer.validated_data["old_password"]):
            return Response({"detail": "Mot de passe actuel incorrect."}, status=status.HTTP_400_BAD_REQUEST)

        user.set_password(serializer.validated_data["new_password"])
        user.save()
        return Response({"detail": "Mot de passe modifié avec succès."})


class UserListView(generics.ListAPIView):
    permission_classes = [IsAdmin]
    serializer_class = UserSerializer
    queryset = User.objects.all()
    search_fields = ["email", "first_name", "last_name", "department"]
    filterset_fields = ["role", "is_active"]
    ordering_fields = ["last_name", "date_joined"]


class UserDetailView(generics.RetrieveUpdateDestroyAPIView):
    permission_classes = [IsAdmin]
    serializer_class = UserSerializer
    queryset = User.objects.all()

    def update(self, request, *args, **kwargs):
        kwargs["partial"] = True
        instance = self.get_object()
        old_data = {"role": instance.role, "is_active": instance.is_active}

        response = super().update(request, *args, **kwargs)

        # Déterminer si c'est un toggle actif/inactif ou une modification générale
        new_data = response.data
        if "is_active" in request.data and len(request.data) == 1:
            action = AuditLog.TOGGLE_ACTIVE
        else:
            action = AuditLog.UPDATE_USER

        log_action(
            user=request.user,
            action=action,
            target_type="user",
            target_id=instance.id,
            target_repr=instance.full_name,
            details={
                "email": instance.email,
                "changes": {k: v for k, v in request.data.items() if k not in ("password",)},
            },
            ip=get_client_ip(request),
        )
        return response

    def destroy(self, request, *args, **kwargs):
        user = self.get_object()
        if user == request.user:
            return Response(
                {"detail": "Vous ne pouvez pas supprimer votre propre compte."},
                status=status.HTTP_400_BAD_REQUEST,
            )
        log_action(
            user=request.user,
            action=AuditLog.DELETE_USER,
            target_type="user",
            target_id=user.id,
            target_repr=user.full_name,
            details={"email": user.email, "role": user.role},
            ip=get_client_ip(request),
        )
        user.delete()
        return Response(status=status.HTTP_204_NO_CONTENT)


class ResendInvitationView(APIView):
    permission_classes = [IsAdmin]

    def post(self, request, pk):
        try:
            user = User.objects.get(pk=pk)
        except User.DoesNotExist:
            return Response({"detail": "Utilisateur introuvable."}, status=status.HTTP_404_NOT_FOUND)

        user.invitation_token = uuid.uuid4()
        user.invitation_sent_at = timezone.now()
        user.save(update_fields=["invitation_token", "invitation_sent_at"])
        send_invitation_email.delay(str(user.id))
        return Response({"detail": "Invitation renvoyée."})


class ImportMembersView(APIView):
    permission_classes = [IsAdmin]

    VALID_ROLES = {
        "eleve": User.ELEVE,
        "étudiant": User.ELEVE,
        "etudiant": User.ELEVE,
        "professeur": User.PROFESSEUR,
        "enseignant": User.PROFESSEUR,
        "pat": User.PAT,
        "personnel_admin": User.PAT,
        "responsable": User.RESPONSABLE,
        "admin": User.ADMIN,
        "administrateur": User.ADMIN,
    }

    def post(self, request):
        file = request.FILES.get("file")
        if not file:
            return Response({"detail": "Fichier Excel requis."}, status=status.HTTP_400_BAD_REQUEST)

        if not file.name.endswith(".xlsx"):
            return Response({"detail": "Format accepté : .xlsx uniquement."}, status=status.HTTP_400_BAD_REQUEST)

        try:
            wb = openpyxl.load_workbook(file)
            ws = wb.active
        except Exception:
            return Response({"detail": "Fichier Excel invalide."}, status=status.HTTP_400_BAD_REQUEST)

        rows_processed = 0
        rows_created = 0
        rows_updated = 0
        rows_errors = 0
        errors = []

        headers = [str(cell.value).strip().lower() if cell.value else "" for cell in ws[1]]
        required = ["nom", "prénom", "email", "rôle"]

        for req in required:
            if req not in headers and req.replace("é", "e").replace("ô", "o") not in headers:
                return Response(
                    {"detail": f"Colonne manquante : '{req}'. Colonnes requises : Nom, Prénom, Email, Rôle, Département, Statut"},
                    status=status.HTTP_400_BAD_REQUEST,
                )

        def get_col(name):
            variants = [name, name.replace("é", "e").replace("ô", "o").replace("è", "e")]
            for v in variants:
                if v in headers:
                    return headers.index(v)
            return None

        col_nom = get_col("nom")
        col_prenom = get_col("prénom") or get_col("prenom")
        col_email = get_col("email")
        col_role = get_col("rôle") or get_col("role")
        col_dept = get_col("département") or get_col("departement")
        col_statut = get_col("statut")

        new_users = []

        for row_idx, row in enumerate(ws.iter_rows(min_row=2, values_only=True), start=2):
            if not any(row):
                continue
            rows_processed += 1

            email = str(row[col_email]).strip().lower() if row[col_email] else ""
            if not email or "@" not in email:
                errors.append({"ligne": row_idx, "erreur": "Email invalide ou manquant."})
                rows_errors += 1
                continue

            last_name = str(row[col_nom]).strip() if row[col_nom] else ""
            first_name = str(row[col_prenom]).strip() if row[col_prenom] else ""
            role_raw = str(row[col_role]).strip().lower() if row[col_role] else "eleve"
            role = self.VALID_ROLES.get(role_raw, User.ELEVE)
            department = str(row[col_dept]).strip() if col_dept is not None and row[col_dept] else ""
            is_active = True
            if col_statut is not None and row[col_statut]:
                is_active = str(row[col_statut]).strip().lower() != "inactif"

            existing = User.objects.filter(email=email).first()
            if existing:
                existing.first_name = first_name
                existing.last_name = last_name
                existing.role = role
                existing.department = department
                existing.is_active = is_active
                existing.save()
                rows_updated += 1
            else:
                user = User(
                    email=email,
                    first_name=first_name,
                    last_name=last_name,
                    role=role,
                    department=department,
                    is_active=is_active,
                    invitation_sent_at=timezone.now(),
                )
                new_users.append(user)
                rows_created += 1

        if new_users:
            User.objects.bulk_create(new_users)
            for user in new_users:
                send_invitation_email.delay(str(user.id))

        import_log = MemberImport.objects.create(
            imported_by=request.user,
            file_name=file.name,
            rows_processed=rows_processed,
            rows_created=rows_created,
            rows_updated=rows_updated,
            rows_errors=rows_errors,
            errors=errors,
        )

        log_action(
            user=request.user,
            action=AuditLog.IMPORT_USERS,
            target_type="import",
            target_repr=file.name,
            details={
                "rows_created": rows_created,
                "rows_updated": rows_updated,
                "rows_errors": rows_errors,
            },
            ip=get_client_ip(request),
        )

        return Response({
            "rows_processed": rows_processed,
            "rows_created": rows_created,
            "rows_updated": rows_updated,
            "rows_errors": rows_errors,
            "errors": errors,
        }, status=status.HTTP_200_OK)


class CreateMemberView(APIView):
    permission_classes = [IsAdmin]

    @staticmethod
    def slugify_name(name):
        import unicodedata
        name = unicodedata.normalize("NFD", name)
        name = "".join(c for c in name if unicodedata.category(c) != "Mn")
        return name.lower().strip().replace(" ", "")

    def post(self, request):
        first_name = request.data.get("first_name", "").strip()
        last_name = request.data.get("last_name", "").strip()
        role = request.data.get("role", "eleve")

        if not first_name or not last_name:
            return Response({"detail": "Nom et prénom obligatoires."}, status=status.HTTP_400_BAD_REQUEST)
        if role not in [User.ELEVE, User.PROFESSEUR, User.PAT, User.RESPONSABLE, User.ADMIN]:
            return Response({"detail": "Rôle invalide."}, status=status.HTTP_400_BAD_REQUEST)

        slug_first = self.slugify_name(first_name)
        slug_last = self.slugify_name(last_name)
        base_email = f"{slug_first}.{slug_last}@ensmg.sn"

        email = base_email
        counter = 2
        while User.objects.filter(email=email).exists():
            email = f"{slug_first}.{slug_last}{counter}@ensmg.sn"
            counter += 1

        user = User.objects.create(
            email=email,
            first_name=first_name,
            last_name=last_name,
            role=role,
            is_active=True,
            password_set=True,
        )
        user.set_password("passer01")
        user.save(update_fields=["password"])

        log_action(
            user=request.user,
            action=AuditLog.CREATE_USER,
            target_type="user",
            target_id=user.id,
            target_repr=user.full_name,
            details={"email": email, "role": role},
            ip=get_client_ip(request),
        )

        data = UserSerializer(user).data
        data["generated_email"] = email
        return Response(data, status=status.HTTP_201_CREATED)


class ImportHistoryView(generics.ListAPIView):
    permission_classes = [IsAdmin]
    serializer_class = MemberImportSerializer
    queryset = MemberImport.objects.all()


class AuditLogView(generics.ListAPIView):
    permission_classes = [IsAuditor]
    serializer_class = AuditLogSerializer
    filterset_fields = ["action", "target_type"]
    search_fields = ["target_repr", "user__first_name", "user__last_name", "user__email"]
    ordering_fields = ["created_at"]
    ordering = ["-created_at"]

    def get_queryset(self):
        return AuditLog.objects.select_related("user").all()
